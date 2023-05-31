import openpyxl

def obtener_links(path):
	file = path
	workbook = openpyxl.load_workbook(file) # Abro el archivo "file" con openpyxl
	hoja_links = workbook["Hoja1"]  # La página con los datos, la primera
	filas = hoja_links.max_row

	links_ingredientes = {}

	for row in range(2,filas+1):   # Agarro el link y lo pongo en "links"
		ingrediente = hoja_links.cell(row=row,column=1).value.lower()
  
		links_coto = hoja_links.cell(row=row,column=2).value
		links_coto = links_coto.split(",") if links_coto else []
  
		links_dia = hoja_links.cell(row=row,column=3).value
		links_dia = links_dia.split(",") if links_dia else []
  
		links_carrefour = hoja_links.cell(row=row,column=4).value
		links_carrefour = links_carrefour.split(",") if links_carrefour else []

		links_ingredientes[ingrediente] = {
									"nombre":ingrediente,
	  								"links_coto":links_coto,
									"links_dia":links_dia,
									"links_carrefour":links_carrefour,
									"html_coto":[],
									"html_dia":[],
									"html_carrefour":[],
									"precios":[],
									"unidades":[],
									"fila":row }

	# for ingrediente in links_ingredientes.values():
	# 	print(ingrediente)
	return links_ingredientes

obtener_links("precios_ingredientes.xlsx")