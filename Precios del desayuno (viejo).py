import time
time1 = time.time()
import requests
from bs4 import BeautifulSoup
import openpyxl
import asyncio
import aiohttp


async def descargar(url):
	async with aiohttp.ClientSession() as session: # Función con asyncio para descargar las páginas
		async with session.get(url) as response:
			text = await response.read()
			return BeautifulSoup(text.decode('utf-8'), 'html.parser') # Devuelve la página ya formateada

def descarga_de_paginas(links):
	loop = asyncio.get_event_loop()			# Inicio Loop de asyncio
	coroutines = [requests.get(i) for i in links]    # Descargo cada link
	lista_de_paginas = loop.run_until_complete(asyncio.gather(*coroutines)) # Cierro el loop con los resultados
	# print(lista_de_paginas)
	return lista_de_paginas

def obtener_links(path):
	file = path
	links = []
	workbook = openpyxl.load_workbook(file) # Abro el archivo "file" con openpyxl
	hoja_links = workbook["Hoja1"]  # La página con los datos, la primera
	filas = hoja_links.max_row  

	for row in range(2,filas+1):   # Agarro el link y lo pongo en "links"
		print("Libro: " + hoja_links.cell(row=row,column=1).value + " //", "Link:" + hoja_links.cell(row=row,column=2).value)
		link = hoja_links.cell(row=row,column=2).value
		links.append(link)

def coto(pagina):
	pagina = requests.get(pagina)
	page = BeautifulSoup(pagina.content, 'html.parser')
	results = page.find("div", id="atg_store_productMoreInfo")
	precio = results.text.split("$")[1].split(",")[0].replace(".", "") # Del resultado, elegir el texto entre "$" y "," que resulta ser el precio, y luego eliminar los puntos
	print("Precio por kilo: $" + precio)
	print()


def dia(pagina):
	pagina = requests.get(pagina)
	page = BeautifulSoup(pagina.content, 'html.parser')
	results = page.find("div", class_="pr0.items-stretch.vtex-flex-layout-0-x-stretchChildrenWidth.flex")
	# print(results)
	texto = page.text.replace("\n","")
	precio = texto.split("$")[-1].split("KG")[0].split("00")[1].split("%")[-1].replace(".","")
	print("Precio por kilo: $" + precio)
	print()


def carrefour(pagina):
	pagina = requests.get(pagina)
	page = BeautifulSoup(pagina.content, 'html.parser')
	results = page.find("div", class_="vtex-store-components-3-x-content h-auto")
	precio = results.text.split("$")[1].split(",")[0].replace(".","")
	print("Precio por kilo: $" + precio)
	print()


coto("https://www.cotodigital3.com.ar/sitios/cdigi/producto/-cafe-instantaneo-nescafe-tradicion-x-100gr/_/A-00282515-00282515-200")
dia("https://diaonline.supermercadosdia.com.ar/nescafe-dolca-suave-ideal-para-batir-100-gr-278969/p")
# dia("https://diaonline.supermercadosdia.com.ar/cafe-instantaneo-nescafe-dolca-mixes-mocca-doypack-125-gr-271407/p")
# vea("https://www.vea.com.ar/cafe-instantaneo-nescafe-tradicion-170-gr/p")
carrefour("https://www.carrefour.com.ar/cafe-instantaneo-nescafe-dolca-suave-doypack-170-g-687505/p")